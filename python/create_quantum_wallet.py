"""
Quantum Wallet Export Generator

Creates an XLSX file compatible with Quantum system import format
from BTG Position API response data.

Usage:
    from create_quantum_wallet import create_quantum_wallet
    from btg_api_position import get_position_by_account, get_position_by_account_and_date

    # Option 1: Get real-time position
    position = get_position_by_account("004209281")
    create_quantum_wallet(position, "25005L01 - Ricardo Camargo Veirano", "output.xlsx")

    # Option 2: Get position at specific date
    position_data = get_position_by_account_and_date("004209281", "2025-12-31")
    create_quantum_wallet(position_data["Position"], "25005L01 - Ricardo Camargo Veirano", "output.xlsx")
"""

from __future__ import annotations

import uuid
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from btg_api_position import (
    Fund,
    Position,
    PositionData,
    Positions,
    get_position_by_account,
    get_position_by_account_and_date,
)


# =============================================================================
# Constants
# =============================================================================

QUANTUM_COLUMNS = [
    "Portfólio",
    "Código do Boleto",
    "Ativo",
    "Tipo de Ativo",
    "Tipo de Movimentação",
    "Data da Movimentação",
    "Data da Conversão",
    "Data da Liquidação",
    "Preço",
    "Quantidade",
    "Valor",
    "Política de Custo",
    "Indexador",
    "Função",
    "Tipo de Taxa",
    "Taxa",
    "Data do Vencimento",
    "Situação do Ativo",
    "Tributação",
    "Gross Up",
    "Risco",
    "Estratégia",
    "Enquadramento",
    "Corretora/Banco",
    "Comentário",
    "Classificação",
    "CNPJ",
    "Benchmark",
    "Gestão",
    "Administrador",
    "Custodiante",
    "Emissor",
    "Setor",
    "Marcação",
    "Fonte",
    "Juros",
    "ISIN",
    "Apelido",
    "Identificador Ativo",
    "Código do Boleto de Aplicação",
    "Adicionar/Retirar Caixa",
    "Excluir",
    "Tipo Liquidez",
    "Liquidez Vencimento",
    "Liquidez Carência",
    "Liquidez D+",
]

# Mapping of TipoCvm to taxation rules
TAXATION_MAP = {
    "2": "Longo Prazo",      # FIRF
    "4": "Longo Prazo",      # FIM
    "6": "FIA/FIP",          # FIA
    "10": "Regra Geral",     # FIDC
    "12": "FIA/FIP",         # FIP
    "13": "FII",             # FII
}

# Risk mapping by asset/fund type
RISK_MAP = {
    "FIA": 5,
    "FIP": 5,
    "FII": 3,
    "FIRF": 1,
    "FIM": 3,
    "FIDC": 4,
    "FixedIncome": 1,
    "Equity": 5,
    "Pension": 1,
}

# Strategy mapping
STRATEGY_MAP = {
    "2": "Renda Fixa - Pós Fixado",
    "4": "Multimercado",
    "6": "Ações - Brasil",
    "10": "Multimercado",
    "12": "Alternativos",
    "13": "Imobiliário",
}

# Manager to Administrator mapping (common administrators)
ADMIN_MAP = {
    "BTG PACTUAL": "BTG Pactual Serviços Financeiros",
    "ITAU": "Intrag DTVM",
    "CREDIT SUISSE": "UBS Brasil Corretora",
    "CSHG": "UBS Brasil Corretora",
    "BNY": "BNY Mellon Serviços Financeiros",
    "SANTANDER": "Santander Caceis",
    "SAFRA": "BTG Pactual Serviços Financeiros",
}

# Manager to Custodian mapping
CUSTODIAN_MAP = {
    "BTG PACTUAL": "Banco BTG Pactual",
    "ITAU": "Itaú Unibanco",
    "CREDIT SUISSE": "Itaú Unibanco",
    "CSHG": "Itaú Unibanco",
    "BNY": "BNY Mellon Banco",
    "SANTANDER": "Santander Caceis",
    "SAFRA": "Banco BTG Pactual",
}


# =============================================================================
# Helper Functions
# =============================================================================

def _safe_float(value: Any, default: float = 0.0) -> float:
    """Safely convert a value to float."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def _format_date(value: Any) -> str:
    """Format date to DD/MM/YYYY format for Quantum."""
    if not value:
        return "Não se Aplica"
    try:
        date_str = str(value)
        if "T" in date_str:
            date_str = date_str.split("T")[0]
        # Parse YYYY-MM-DD and convert to DD/MM/YYYY
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return "Não se Aplica"


def _format_currency(value: float) -> float:
    """Return numeric value for XLSX export (Quantum expects numbers, not strings)."""
    return value


def _get_administrator(manager_name: str) -> str:
    """Get administrator based on manager name."""
    if not manager_name:
        return "Outros"
    manager_upper = manager_name.upper()
    for key, admin in ADMIN_MAP.items():
        if key in manager_upper:
            return admin
    return "Outros"


def _get_custodian(manager_name: str) -> str:
    """Get custodian based on manager name."""
    if not manager_name:
        return "Outros"
    manager_upper = manager_name.upper()
    for key, custodian in CUSTODIAN_MAP.items():
        if key in manager_upper:
            return custodian
    return "Outros"


def _get_taxation(tipo_cvm: str, fund_name: str = "") -> str:
    """Get taxation regime based on CVM type."""
    # Check for special cases in fund name
    fund_upper = fund_name.upper() if fund_name else ""
    if "PREV" in fund_upper or "VGBL" in fund_upper or "PGBL" in fund_upper:
        return "Isento (IR + IOF)"
    if "INFRA" in fund_upper and ("FIRF" in fund_upper or "RF" in fund_upper):
        return "Isento (IR + IOF)"
    return TAXATION_MAP.get(tipo_cvm, "Longo Prazo")


def _get_risk(tipo_cvm: str, benchmark: str | None = None) -> int:
    """Get risk level (1-5) based on CVM type."""
    # FIA/FIP = 5, FII = 3, FIM = 3, FIRF = 1, FIDC = 4
    risk_by_cvm = {
        "2": 1,   # FIRF
        "4": 3,   # FIM
        "6": 5,   # FIA
        "10": 4,  # FIDC
        "12": 3,  # FIP (Alternativos)
        "13": 3,  # FII
    }
    return risk_by_cvm.get(tipo_cvm, 3)


def _get_strategy(tipo_cvm: str, benchmark: str | None = None, fund_name: str = "") -> str:
    """Get strategy based on CVM type and benchmark."""
    fund_upper = fund_name.upper() if fund_name else ""

    # Check for specific strategies in fund name
    if "CRYPTO" in fund_upper or "HASH" in fund_upper:
        return "Alternativos"
    if "INFRA" in fund_upper:
        return "Renda Fixa - Diversos"
    if "CREDIT" in fund_upper or "CRÉDITO" in fund_upper:
        return "Multimercado"
    if benchmark == "IPCA":
        return "Renda Fixa - Inflação"
    if benchmark == "CDI":
        if tipo_cvm in ["2"]:
            return "Renda Fixa - Pós Fixado"
        return "Multimercado"
    if benchmark == "IBOVESPA" or tipo_cvm == "6":
        return "Ações - Brasil"

    return STRATEGY_MAP.get(tipo_cvm, "Multimercado")


def _get_enquadramento(tipo_cvm: str, asset_type: str = "") -> str:
    """Get enquadramento (classification) for asset."""
    if asset_type in ["FixedIncome", "RF"]:
        return "Renda Fixa Público" if "BACEN" in asset_type or "NTNB" in asset_type or "LTN" in asset_type else "Renda Fixa Privado"
    if asset_type in ["Equity", "FII"]:
        return "Renda Variável"
    return "Fundos"


def _generate_boleto_code(
    portfolio: str,
    ativo: str,
    data: str,
    quantidade: float
) -> str:
    """Generate a deterministic boleto code (UUID) based on position key fields.

    Uses UUID5 with a namespace to create a reproducible UUID from the input fields.
    The same inputs will always generate the same UUID.

    Args:
        portfolio: Portfolio name
        ativo: Asset name
        data: Position date (DD/MM/YYYY format)
        quantidade: Quantity of shares/units

    Returns:
        Deterministic UUID string
    """
    # Create a unique key from the position fields
    key = f"{portfolio}|{ativo}|{data}|{quantidade:.6f}"

    # Use UUID5 with DNS namespace to generate deterministic UUID
    return str(uuid.uuid5(uuid.NAMESPACE_DNS, key))


# =============================================================================
# Parsing Functions
# =============================================================================

def _parse_investment_funds(data: Position, portfolio: str) -> list[dict[str, Any]]:
    """Parse InvestmentFund positions for Quantum format.

    Creates one consolidated row per fund with:
    - Quantidade = sum of NumberOfShares across all acquisitions
    - Valor = sum of GrossAssetValue across all acquisitions
    - Preço = total GrossAssetValue / total NumberOfShares (current price per share)
    """
    funds = data.get("InvestmentFund", []) or []
    rows = []

    for fund in funds:
        fund_info = fund.get("Fund", {})
        acquisitions = fund.get("Acquisition", []) or []

        fund_name = fund_info.get("FundName", "")
        cnpj = fund_info.get("FundCNPJCode", "")
        manager = fund_info.get("ManagerName", "")
        benchmark = fund_info.get("BenchMark")
        tipo_cvm = fund_info.get("TipoCvm", "4")
        liquidity = fund_info.get("FundLiquidity", "0")

        # Sum values across all acquisitions
        total_gross_value = sum(_safe_float(acq.get("GrossAssetValue")) for acq in acquisitions)
        total_shares = sum(_safe_float(acq.get("NumberOfShares")) for acq in acquisitions)

        # Calculate price as total GrossAssetValue / total NumberOfShares
        price = total_gross_value / total_shares if total_shares > 0 else 0.0

        # Determine asset type based on TipoCvm
        tipo_ativo_map = {
            "2": "FI",    # FIRF
            "4": "FI",    # FIM
            "6": "FI",    # FIA
            "10": "FIDC",  # FIDC
            "12": "FIP",  # FIP
            "13": "FII",  # FII
        }
        tipo_ativo = tipo_ativo_map.get(tipo_cvm, "FI")

        # Get position date from data
        position_date = _format_date(data.get("PositionDate"))

        # Generate deterministic boleto code
        boleto_code = _generate_boleto_code(portfolio, fund_name, position_date, total_shares)

        row = {
            "Portfólio": portfolio,
            "Código do Boleto": boleto_code,
            "Ativo": fund_name,
            "Tipo de Ativo": tipo_ativo,
            "Tipo de Movimentação": "Aplicação",
            "Data da Movimentação": position_date,
            "Data da Conversão": position_date,
            "Data da Liquidação": position_date,
            "Preço": _format_currency(price),
            "Quantidade": total_shares,
            "Valor": _format_currency(total_gross_value),
            "Política de Custo": "Não Possui",
            "Indexador": "Não se Aplica",
            "Função": "Não se Aplica ",
            "Tipo de Taxa": "Não se Aplica",
            "Taxa": "Não se Aplica",
            "Data do Vencimento": "Não se Aplica ",
            "Situação do Ativo": "Em funcionamento normal",
            "Tributação": _get_taxation(tipo_cvm, fund_name),
            "Gross Up": "Não se Aplica ",
            "Risco": _get_risk(tipo_cvm, benchmark),
            "Estratégia": _get_strategy(tipo_cvm, benchmark, fund_name),
            "Enquadramento": "Fundos",
            "Corretora/Banco": "BTG Pactual",
            "Comentário": "Não Informado",
            "Classificação": _get_classification(tipo_cvm, benchmark, fund_name),
            "CNPJ": _format_cnpj(cnpj),
            "Benchmark": benchmark if benchmark else "Não Informado",
            "Gestão": _format_manager(manager),
            "Administrador": _get_administrator(manager),
            "Custodiante": _get_custodian(manager),
            "Emissor": "Outros",
            "Setor": "Outros",
            "Marcação": "Não se Aplica",
            "Fonte": "Padrão",
            "Juros": "Não se Aplica",
            "ISIN": _get_isin_from_fund(fund_info),
            "Apelido": "Não Informado",
            "Identificador Ativo": "Não se Aplica",
            "Código do Boleto de Aplicação": "",
            "Adicionar/Retirar Caixa": "Não",
            "Excluir": "",
            "Tipo Liquidez": "D+",
            "Liquidez Vencimento": "Não se Aplica",
            "Liquidez Carência": "Não se Aplica",
            "Liquidez D+": "Não se Aplica",
        }
        rows.append(row)

    return rows


def _parse_fixed_income(data: Position, portfolio: str) -> list[dict[str, Any]]:
    """Parse FixedIncome positions for Quantum format.

    Creates one consolidated row per fixed income asset with:
    - Quantidade = total Quantity
    - Valor = GrossValue (current market value)
    - Preço = GrossValue / Quantity (current price per unit)
    """
    fixed_income = data.get("FixedIncome", []) or []
    rows = []

    # Get position date from data
    position_date = _format_date(data.get("PositionDate"))

    for asset in fixed_income:
        ticker = asset.get("Ticker", "")
        issuer = asset.get("Issuer", "")
        index_name = asset.get("ReferenceIndexName", "")
        index_rate = asset.get("IndexYieldRate", "")
        maturity = _format_date(asset.get("MaturityDate"))
        isin = asset.get("ISIN", "")
        issuer_type = asset.get("IssuerType", "")
        accounting_group = asset.get("AccountingGroupCode", "")

        # Get consolidated position values
        quantity = _safe_float(asset.get("Quantity"))
        gross_value = _safe_float(asset.get("GrossValue"))

        # Calculate price as GrossValue / Quantity
        price = gross_value / quantity if quantity > 0 else 0.0

        # Determine tipo de ativo based on accounting group
        tipo_ativo = "Títulos Públicos Líquidos" if issuer_type == "Titulo Publico" else accounting_group

        # Determine enquadramento
        enquadramento = "Renda Fixa Público" if issuer_type == "Titulo Publico" else "Renda Fixa Privado"

        # Determine strategy based on index
        if index_name == "IPCA":
            estrategia = "Renda Fixa - Inflação"
            classificacao = "Renda Fixa - Inflação"
        elif index_name == "PRE":
            estrategia = "Renda Fixa - Prefixado"
            classificacao = "Renda Fixa - Prefixado"
        else:
            estrategia = "Renda Fixa - Pós Fixado"
            classificacao = "Renda Fixa - Pós Fixado"

        asset_name = f"{ticker} - {issuer}" if ticker and issuer else ticker or issuer
        boleto_code = _generate_boleto_code(portfolio, asset_name, position_date, quantity)

        row = {
            "Portfólio": portfolio,
            "Código do Boleto": boleto_code,
            "Ativo": asset_name,
            "Tipo de Ativo": tipo_ativo,
            "Tipo de Movimentação": "Aplicação",
            "Data da Movimentação": position_date,
            "Data da Conversão": position_date,
            "Data da Liquidação": position_date,
            "Preço": _format_currency(price),
            "Quantidade": quantity,
            "Valor": _format_currency(gross_value),
            "Política de Custo": "Não Possui",
            "Indexador": index_name if index_name else "Não se Aplica",
            "Função": "Não se Aplica ",
            "Tipo de Taxa": "Não se Aplica",
            "Taxa": index_rate if index_rate else "Não se Aplica",
            "Data do Vencimento": maturity,
            "Situação do Ativo": "Ativo",
            "Tributação": "Isento (IR + IOF)" if accounting_group == "CRI" else "Regra Geral",
            "Gross Up": "Não se Aplica ",
            "Risco": 1,
            "Estratégia": estrategia,
            "Enquadramento": enquadramento,
            "Corretora/Banco": "BTG Pactual",
            "Comentário": "Não Informado",
            "Classificação": classificacao,
            "CNPJ": "Não se Aplica",
            "Benchmark": "CDI",
            "Gestão": "Outros",
            "Administrador": "Outros",
            "Custodiante": "Outros",
            "Emissor": issuer if issuer else "TESOURO NACIONAL",
            "Setor": "Outros",
            "Marcação": "Mercado",
            "Fonte": "Padrão",
            "Juros": "Não se Aplica",
            "ISIN": isin if isin else "Não se Aplica",
            "Apelido": f"{ticker} {maturity.split('/')[-1] if '/' in maturity else ''}" if ticker else "Não Informado",
            "Identificador Ativo": "Não se Aplica",
            "Código do Boleto de Aplicação": "",
            "Adicionar/Retirar Caixa": "Não",
            "Excluir": "",
            "Tipo Liquidez": "Vencimento",
            "Liquidez Vencimento": "Não se Aplica",
            "Liquidez Carência": "Não se Aplica",
            "Liquidez D+": "Não se Aplica",
        }
        rows.append(row)

    return rows


def _parse_pension(data: Position, portfolio: str) -> list[dict[str, Any]]:
    """Parse PensionInformations positions for Quantum format.

    Creates one consolidated row per pension position with:
    - Quantidade = NumberOfShares
    - Valor = GrossAssetValue (current value)
    - Preço = GrossAssetValue / NumberOfShares (current price per share)
    """
    pensions = data.get("PensionInformations", []) or []
    rows = []

    # Get position date from data
    position_date = _format_date(data.get("PositionDate"))

    for pension in pensions:
        fund_type = pension.get("FundType", "VGBL")

        positions = pension.get("Positions", []) or []

        for pos in positions:
            fund_name = pos.get("FundName", "")
            cnpj = pos.get("PensionCnpjCode", "")

            # Get consolidated position values
            number_of_shares = _safe_float(pos.get("NumberOfShares"))
            gross_asset_value = _safe_float(pos.get("GrossAssetValue"))

            # Calculate price as GrossAssetValue / NumberOfShares
            price = gross_asset_value / number_of_shares if number_of_shares > 0 else 0.0

            boleto_code = _generate_boleto_code(portfolio, fund_name, position_date, number_of_shares)

            row = {
                "Portfólio": portfolio,
                "Código do Boleto": boleto_code,
                "Ativo": fund_name,
                "Tipo de Ativo": "FI",
                "Tipo de Movimentação": "Aplicação",
                "Data da Movimentação": position_date,
                "Data da Conversão": position_date,
                "Data da Liquidação": position_date,
                "Preço": _format_currency(price),
                "Quantidade": number_of_shares,
                "Valor": _format_currency(gross_asset_value),
                "Política de Custo": "Não Possui",
                "Indexador": "Não se Aplica",
                "Função": "Não se Aplica ",
                "Tipo de Taxa": "Não se Aplica",
                "Taxa": "Não se Aplica",
                "Data do Vencimento": "Não se Aplica ",
                "Situação do Ativo": "Em funcionamento normal",
                "Tributação": "Isento (IR + IOF)",
                "Gross Up": "Não se Aplica ",
                "Risco": 1,
                "Estratégia": "Renda Fixa - Pós Fixado",
                "Enquadramento": "Fundos",
                "Corretora/Banco": "BTG Pactual",
                "Comentário": "Não Informado",
                "Classificação": "Pós-fixado",
                "CNPJ": _format_cnpj(cnpj),
                "Benchmark": "CDI",
                "Gestão": "BTG Pactual Asset Management",
                "Administrador": "BTG Pactual Serviços Financeiros",
                "Custodiante": "Banco BTG Pactual",
                "Emissor": "Outros",
                "Setor": "Outros",
                "Marcação": "Não se Aplica",
                "Fonte": "Padrão",
                "Juros": "Não se Aplica",
                "ISIN": _get_isin_from_pension(pos),
                "Apelido": f"{fund_type} - {fund_name}",
                "Identificador Ativo": "Não se Aplica",
                "Código do Boleto de Aplicação": "",
                "Adicionar/Retirar Caixa": "Não",
                "Excluir": "",
                "Tipo Liquidez": "D+",
                "Liquidez Vencimento": "Não se Aplica",
                "Liquidez Carência": "Não se Aplica",
                "Liquidez D+": "Não se Aplica",
            }
            rows.append(row)

    return rows


def _parse_equities(data: Position, portfolio: str) -> list[dict[str, Any]]:
    """Parse Equities positions for Quantum format.

    Creates one consolidated row per equity position with:
    - Quantidade = Quantity
    - Valor = GrossValue (current market value)
    - Preço = MarketPrice (current market price per share)
    """
    equities = data.get("Equities", []) or []
    rows = []

    # Get position date from data
    position_date = _format_date(data.get("PositionDate"))

    for equity in equities:
        stocks = equity.get("StockPositions", []) or []

        for stock in stocks:
            ticker = stock.get("Ticker", "")
            description = stock.get("Description", "")
            isin = stock.get("ISINCode", "")
            is_fii = stock.get("IsFII", "false") == "true"
            sector = stock.get("SectorDescription", "")

            # Get consolidated position values
            quantity = _safe_float(stock.get("Quantity"))
            gross_value = _safe_float(stock.get("GrossValue"))
            market_price = _safe_float(stock.get("MarketPrice"))

            # Use market price, or calculate from GrossValue/Quantity
            price = market_price if market_price > 0 else (gross_value / quantity if quantity > 0 else 0.0)

            boleto_code = _generate_boleto_code(portfolio, description, position_date, quantity)

            tipo_ativo = "Ação"
            tributacao = "FII" if is_fii else "Ações"
            estrategia = "Alternativos" if is_fii else "Alternativos"
            risco = 3 if is_fii else 5

            row = {
                "Portfólio": portfolio,
                "Código do Boleto": boleto_code,
                "Ativo": description,
                "Tipo de Ativo": tipo_ativo,
                "Tipo de Movimentação": "Aplicação",
                "Data da Movimentação": position_date,
                "Data da Conversão": position_date,
                "Data da Liquidação": position_date,
                "Preço": _format_currency(price),
                "Quantidade": quantity,
                "Valor": _format_currency(gross_value),
                "Política de Custo": "Não Possui",
                "Indexador": "Não se Aplica",
                "Função": "Não se Aplica ",
                "Tipo de Taxa": "Não se Aplica",
                "Taxa": "Não se Aplica",
                "Data do Vencimento": "Não se Aplica ",
                "Situação do Ativo": "Ativo",
                "Tributação": tributacao,
                "Gross Up": "Não se Aplica ",
                "Risco": risco,
                "Estratégia": estrategia,
                "Enquadramento": "Renda Variável",
                "Corretora/Banco": "BTG Pactual",
                "Comentário": "Não Informado",
                "Classificação": "Outros",
                "CNPJ": "Não se Aplica",
                "Benchmark": "Ibovespa",
                "Gestão": "Outros",
                "Administrador": "Outros",
                "Custodiante": "Outros",
                "Emissor": "Não Informado",
                "Setor": "Outros",
                "Marcação": "Não se Aplica",
                "Fonte": "Padrão",
                "Juros": "Não se Aplica",
                "ISIN": isin if isin else "Não se Aplica",
                "Apelido": "Não Informado",
                "Identificador Ativo": "Não se Aplica",
                "Código do Boleto de Aplicação": "",
                "Adicionar/Retirar Caixa": "Não",
                "Excluir": "",
                "Tipo Liquidez": "D+",
                "Liquidez Vencimento": "Não se Aplica",
                "Liquidez Carência": "Não se Aplica",
                "Liquidez D+": "Não se Aplica",
            }
            rows.append(row)

    return rows


# =============================================================================
# Utility Functions
# =============================================================================

def _format_cnpj(cnpj: str) -> str:
    """Format CNPJ with mask XX.XXX.XXX/XXXX-XX."""
    if not cnpj:
        return "Não se Aplica"
    # Remove any existing formatting
    cnpj_clean = "".join(filter(str.isdigit, cnpj))
    if len(cnpj_clean) != 14:
        return cnpj
    return f"{cnpj_clean[:2]}.{cnpj_clean[2:5]}.{cnpj_clean[5:8]}/{cnpj_clean[8:12]}-{cnpj_clean[12:14]}"


def _format_manager(manager: str) -> str:
    """Format manager name for Quantum."""
    if not manager:
        return "Outros"
    # Simplify common manager names
    manager_map = {
        "BTG PACTUAL ASSET MANAGEMENT": "BTG Pactual Asset Management",
        "ITAU UNIBANCO ASSET MANAGEMENT": "Itaú Asset Management",
        "CREDIT SUISSE HEDGING-GRIFFO": "Credit Suisse Hedging-Griffo",
        "SPX GESTAO DE RECURSOS": "SPX Capital",
        "KAPITALO INVESTIMENTOS": "Kapitalo Investimentos",
        "ABSOLUTE GESTAO": "Absolute Investimentos",
        "HASHDEX GESTORA": "Hashdex",
        "GAMA INVESTIMENTOS": "Gama Investimentos",
        "GENOA CAPITAL": "Genoa Capital",
        "ATMOS CAPITAL": "Atmos Capital",
        "SQUADRA INVESTIMENTOS": "Squadra Investimentos",
        "SHARP CAPITAL": "Sharp Capital",
        "ANGA": "Angá Investimentos",
        "AUGME CAPITAL": "Augme Capital",
        "LEGEND WM": "Legend Wealth Management",
        "QUADRA GESTAO": "Quadra Capital",
        "ORRAM GESTAO": "Orram Investimentos",
        "EST GESTAO": "Est Gestão de Patrimônio",
        "GTI ADMINISTRACAO": "GTI Administração de Recursos",
        "SPECTRA INVESTIMENTOS": "Spectra Investimentos",
        "SAFRA ASSET": "Safra Asset",
        "JIVE INVESTMENTS": "Jive Investments",
        "VINCI PARTNERS": "Vinci Partners",
        "KINEA INVESTIMENTOS": "Kinea Investimentos",
    }

    manager_upper = manager.upper()
    for key, value in manager_map.items():
        if key in manager_upper:
            return value
    return manager


def _get_isin_from_fund(fund_info: Fund) -> str:
    """Extract ISIN code from fund info if available."""
    # ISIN is typically not in the fund info, would need separate lookup
    return "Não se Aplica"


def _get_isin_from_pension(pos: Positions) -> str:
    """Extract ISIN code from pension position if available."""
    return "Não se Aplica"


def _get_classification(tipo_cvm: str, benchmark: str | None, fund_name: str) -> str:
    """Get classification based on fund characteristics."""
    fund_upper = fund_name.upper() if fund_name else ""

    if "CRYPTO" in fund_upper or "HASH" in fund_upper:
        return "Outros"
    if "INFRA" in fund_upper:
        return "Pós-fixado"
    if benchmark == "IPCA":
        return "Inflação"
    if benchmark == "CDI":
        return "Pós-fixado"
    if benchmark == "IBOVESPA" or tipo_cvm == "6":
        return "Ações"
    if tipo_cvm in ["4", "10"]:
        return "Multimercado"
    if tipo_cvm == "12":
        return "Alternativo"
    if tipo_cvm == "13":
        return "Investimento Imobiliário"

    return "Outros"


# =============================================================================
# Main Functions
# =============================================================================

def create_quantum_wallet(
    position: Position | PositionData,
    portfolio_name: str,
    output_file: str = "quantum_wallet.xlsx"
) -> list[dict[str, Any]]:
    """
    Create a Quantum-compatible XLSX file from BTG Position API response.

    Args:
        position: Position from get_position_by_account() or PositionData from get_position_by_account_and_date()
        portfolio_name: The portfolio identifier (e.g., "25005L01 - Ricardo Camargo Veirano")
        output_file: Output XLSX file path

    Returns:
        List of dictionaries with Quantum-formatted data
    """
    # Handle both Position and PositionData (which wraps Position)
    if "Position" in position:
        data: Position = position["Position"]  # type: ignore[typeddict-item]
    else:
        data = position  # type: ignore[assignment]

    all_rows: list[dict[str, Any]] = []

    # Parse each asset type
    all_rows.extend(_parse_investment_funds(data, portfolio_name))
    all_rows.extend(_parse_fixed_income(data, portfolio_name))
    all_rows.extend(_parse_pension(data, portfolio_name))
    all_rows.extend(_parse_equities(data, portfolio_name))

    # Write to XLSX (Quantum format)
    wb = Workbook()
    ws = wb.create_sheet("Quantum Import", 0)

    # Remove default sheet if it exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Write header row
    ws.append(QUANTUM_COLUMNS)

    # Write data rows
    for row in all_rows:
        ws.append([row.get(col, "") for col in QUANTUM_COLUMNS])

    wb.save(output_file)

    print(f"Exported {len(all_rows)} rows to {output_file}")
    return all_rows


def create_quantum_wallet_from_account(
    account_number: str,
    position_date: str | None,
    portfolio_name: str,
    output_file: str = "quantum_wallet.xlsx"
) -> list[dict[str, Any]]:
    """
    Fetch BTG position data and create Quantum-compatible XLSX.

    Args:
        account_number: BTG account number (e.g., "004209281")
        position_date: Position date in YYYY-MM-DD format (None for real-time)
        portfolio_name: The portfolio identifier for Quantum
        output_file: Output XLSX file path

    Returns:
        List of dictionaries with Quantum-formatted data
    """
    if position_date:
        position_data = get_position_by_account_and_date(account_number, position_date)
        return create_quantum_wallet(position_data, portfolio_name, output_file)
    else:
        position = get_position_by_account(account_number)
        return create_quantum_wallet(position, portfolio_name, output_file)


# =============================================================================
# Entry Point
# =============================================================================

if __name__ == "__main__":
    # Fetch position data from API
    position_data = get_position_by_account_and_date("004209281", "2025-11-28")

    # Create Quantum wallet
    portfolio_name = "25005L01T - Ricardo Camargo Veirano"
    rows = create_quantum_wallet(position_data, portfolio_name, f"./export_data/{portfolio_name}.xlsx")

    print(f"\nGenerated {len(rows)} boletos for Quantum import")
    print("\nSample row:")
    if rows:
        for key, value in list(rows[0].items())[:10]:
            print(f"  {key}: {value}")
